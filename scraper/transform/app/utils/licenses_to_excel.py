import json
import logging
import os

import pandas as pd
from openpyxl.styles import Font, PatternFill
from app.constants import BASE_PDF_URL

logger = logging.getLogger(__name__)


def json_excel(json_path: str, excel_path: str):
    """
    Reads a license JSON file, converts it to a pandas DataFrame,
    sorts by business_name and zipcode, and saves it as an Excel file
    with auto column width and a highlighted header row.
    """
    if not os.path.exists(json_path):
        logger.error(f"Error: {json_path} does not exist.")
        return

    try:
        # Read JSON data
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)

        # Convert to DataFrame
        df = pd.DataFrame(data)
    
        # Build sort config only for columns that exist
        sort_columns = []
        ascending = []
        keys = []

        if "minutes_date" in df.columns:
            sort_columns.append("minutes_date")
            ascending.append(False)   # largest → smallest
            keys.append(None)

        if "business_name" in df.columns:
            sort_columns.append("business_name")
            ascending.append(True)    # smallest → largest
            keys.append(lambda s: s.str.lower())

        if sort_columns:
            df = df.sort_values(
                by=sort_columns,
                ascending=ascending,
                key=lambda col: col.str.lower() if col.name == "business_name" else col,
            )


        # Save to Excel
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Licenses")

            # Access the openpyxl objects
            # workbook = writer.book # (Used implicitly by sheets)
            worksheet = writer.sheets["Licenses"]

            # Convert file_name column to clickable Excel hyperlinks
            if "file_name" in df.columns:
                file_col_idx = df.columns.get_loc("file_name") + 1
                hyperlink_font = Font(color="0000FF", underline="single")
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=file_col_idx)
                    filename = cell.value
                    if filename:
                        url = f"{BASE_PDF_URL}{filename}"
                        cell.value = filename
                        cell.hyperlink = url
                        cell.font = hyperlink_font

            # Auto column width
            for idx, col in enumerate(df.columns):
                # Calculate max length in column
                # series = df[col] # (Not used directly, map below is better)
                max_len = df[col].fillna("").astype(str).map(len).max()
                # Also consider header length
                max_len = max(max_len, len(str(col))) + 2

                # Column letters: A=1, B=2, etc.
                col_letter = worksheet.cell(row=1, column=idx + 1).column_letter
                worksheet.column_dimensions[col_letter].width = min(
                    max_len, 100
                )  # Cap width at 100

            # Highlight header row
            header_font = Font(bold=True, color="FFFFFF")
            # Using a professional blue color for the header background
            header_fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )

            for cell in worksheet[1]:  # Row 1 is the header
                cell.font = header_font
                cell.fill = header_fill

        logger.info(f"Successfully exported data to {excel_path}")
        return True

    except Exception as e:
        # logger.exception("An error occurred")
        logger.error(f"An error occurred: {e}")
        return False


if __name__ == "__main__":
    # Default file paths
    input_json = os.path.join(os.path.dirname(__file__), "..", "all_licenses.json")
    output_xlsx = os.path.join(os.path.dirname(__file__), "..", "licenses.xlsx")

    # Check if all_licenses.json exists relative to this script
    if not os.path.exists(input_json):
        # Try current directory as fallback
        input_json = "all_licenses.json"
        output_xlsx = "licenses.xlsx"

    json_excel(input_json, output_xlsx)
